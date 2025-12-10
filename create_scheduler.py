import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

def create_schedule_template(filename="2025_학사일정_양식.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2025 학사일정"

    # --- 스타일 정의 ---
    # 색상 (이미지 참고: 헤더 파랑/노랑, 일요일 빨강 등)
    color_header_blue = "D9E1F2"  # 연한 파랑
    color_header_yellow = "FFF2CC" # 연한 노랑
    color_header_green = "E2EFDA" # 연한 초록
    
    border_thin = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    font_bold = Font(bold=True, name='맑은 고딕', size=10)
    font_normal = Font(name='맑은 고딕', size=10)

    # --- 헤더 구조 생성 (Row 1, 2) ---
    
    # 1행: 대분류
    headers_row1 = [
        ("날짜", "A", "C"),
        ("출석코드", "D", "D"),
        ("강의일수", "E", "E"),
        ("수업일수", "F", "F"),
        ("창의적체험활동 시수", "G", "L"), # 1,2,3학년
        ("창의적체험활동", "M", "M"),
        ("일정", "N", "P")
    ]
    
    for title, start_col, end_col in headers_row1:
        # 병합
        ws.merge_cells(f"{start_col}1:{end_col}1")
        cell = ws[f"{start_col}1"]
        cell.value = title
        cell.font = font_bold
        cell.alignment = align_center
        cell.border = border_thin
        # 배경색 (구역별로 다르게)
        if title == "날짜": cell.fill = PatternFill(start_color=color_header_blue, fill_type="solid")
        elif title == "일정": cell.fill = PatternFill(start_color=color_header_blue, fill_type="solid")
        elif "체험활동" in title: cell.fill = PatternFill(start_color=color_header_green, fill_type="solid")
        else: cell.fill = PatternFill(start_color=color_header_yellow, fill_type="solid")

        # 병합된 셀 테두리 적용을 위해 순회
        start_idx = openpyxl.utils.column_index_from_string(start_col)
        end_idx = openpyxl.utils.column_index_from_string(end_col)
        for c in range(start_idx, end_idx + 1):
            ws.cell(row=1, column=c).border = border_thin

    # 2행: 상세 분류
    # A~C: 월, 일, 요일
    # D: 출석코드
    # E: 1, F: 2 (강의/수업일수) -- 임의로 설정
    # G~H: 1학년(자/동), I~J: 2학년(자/동), K~L: 3학년(자/동)
    # M: 세부사항
    # N: 학사, O: 교사, P: 행사
    
    headers_row2 = [
        ("월", "A"), ("일", "B"), ("요일", "C"),
        ("출석코드", "D"),
        ("1", "E"), ("2", "F"), # 강의/수업
        ("자", "G"), ("동", "H"), # 1학년
        ("자", "I"), ("동", "J"), # 2학년
        ("자", "K"), ("동", "L"), # 3학년
        ("세부사항", "M"),
        ("학사", "N"), ("교사", "O"), ("행사(교과, 학년, 학생)", "P")
    ]

    for title, col in headers_row2:
        cell = ws[f"{col}2"]
        cell.value = title
        cell.font = font_bold
        cell.alignment = align_center
        cell.border = border_thin
        
        # 배경색 상속 (1행과 비슷하게 맞춤)
        if col in ['A','B','C', 'N','O','P']: 
            cell.fill = PatternFill(start_color=color_header_blue, fill_type="solid")
        elif col in ['G','H','I','J','K','L', 'M']:
            cell.fill = PatternFill(start_color=color_header_green, fill_type="solid")
        else:
            cell.fill = PatternFill(start_color=color_header_yellow, fill_type="solid")

    # 상단 1학년, 2학년, 3학년 구분 행 추가 (1.5행 느낌으로 만들어야 하지만 엑셀은 불가하니 2행 헤더를 조금 수정하거나 그대로 둠)
    # 이미지상으로는 '참의적체험활동 시수' 아래에 '1학년', '2학년', '3학년'이 있고 그 아래 '자', '동'이 있는 3단 구조임.
    # 복잡하므로 1행을 늘리거나 행을 추가해야 함.
    # -> 행을 하나 추가하자.
    
    ws.insert_rows(2) # 2행에 새 행 삽입 -> 기존 2행은 3행이 됨
    
    # 2행 (중간 헤더: 학년 구분 등)
    mid_headers = [
        ("1학년", "G", "H"), ("2학년", "I", "J"), ("3학년", "K", "L")
    ]
    for title, start_col, end_col in mid_headers:
        ws.merge_cells(f"{start_col}2:{end_col}2")
        cell = ws[f"{start_col}2"]
        cell.value = title
        cell.font = font_bold
        cell.alignment = align_center
        cell.border = border_thin
        cell.fill = PatternFill(start_color=color_header_green, fill_type="solid")
        
        start_idx = openpyxl.utils.column_index_from_string(start_col)
        end_idx = openpyxl.utils.column_index_from_string(end_col)
        for c in range(start_idx, end_idx + 1):
            ws.cell(row=2, column=c).border = border_thin
            
    # 나머지 빈칸 병합 (날짜 등은 1~2행 병합 필요)
    for col in ['A','B','C','D','E','F','M','N','O','P']:
        ws.merge_cells(f"{col}1:{col}2")
    
    # -----------------------------
    # 데이터 행 예시 (빈칸 50개 생성)
    # -----------------------------
    start_row = 4
    for i in range(50):
        r = start_row + i
        for col_idx in range(1, 17): # A ~ P
            cell = ws.cell(row=r, column=col_idx)
            cell.border = border_thin
            cell.alignment = align_center
            cell.font = font_normal
            
            # 요일(C열)이 '토', '일'이면 빨간색 글씨
            if col_idx == 3: # C열
                pass # 나중에 데이터 넣을때 처리
            
            # 너비 조정
            col_letter = get_column_letter(col_idx)
            if col_letter in ['A', 'B', 'C']: ws.column_dimensions[col_letter].width = 5
            elif col_letter == 'M': ws.column_dimensions[col_letter].width = 20
            elif col_letter in ['N', 'O', 'P']: ws.column_dimensions[col_letter].width = 30
            else: ws.column_dimensions[col_letter].width = 8

    # 틀 고정 (3행까지 고정)
    ws.freeze_panes = "D4"

    wb.save(filename)
    print(f"엑셀 파일 생성 완료: {filename}")

if __name__ == "__main__":
    create_schedule_template()
